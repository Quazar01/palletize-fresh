import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SPECIAL_SHEETS = {"ICA Helg", "Axfood Pontus", "Axfood Pontus BIGGER"}


def is_non_empty(value):
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def text_value(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(value):
    return text_value(value).lower()


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def is_date_like(value):
    if isinstance(value, (date, datetime)):
        return True
    if isinstance(value, str):
        v = value.strip()
        return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", v))
    return False


def parse_customer_name(kund_cell_value):
    raw = text_value(kund_cell_value)
    if ":" not in raw:
        return raw
    return raw.split(":", 1)[1].strip()


def a1(row, column):
    return f"{get_column_letter(column)}{row}"


def range_a1(min_row, min_col, max_row, max_col):
    return f"{a1(min_row, min_col)}:{a1(max_row, max_col)}"


def scan_sheet_cells(worksheet):
    non_empty = {}
    formulas = []
    lower_text_index = []

    for cell in worksheet._cells.values():
        if not is_non_empty(cell.value):
            continue
        key = (cell.row, cell.column)
        non_empty[key] = cell.value
        if is_formula(cell.value):
            formulas.append((cell.row, cell.column, cell.value))
        if isinstance(cell.value, str):
            lower_text_index.append((cell.row, cell.column, normalize_text(cell.value), cell.value))

    all_coords = list(non_empty.keys())
    if all_coords:
        used_min_row = min(r for r, _ in all_coords)
        used_max_row = max(r for r, _ in all_coords)
        used_min_col = min(c for _, c in all_coords)
        used_max_col = max(c for _, c in all_coords)
    else:
        used_min_row = used_max_row = 1
        used_min_col = used_max_col = 1

    return {
        "values": non_empty,
        "formulas": formulas,
        "textIndex": lower_text_index,
        "usedMinRow": used_min_row,
        "usedMaxRow": used_max_row,
        "usedMinCol": used_min_col,
        "usedMaxCol": used_max_col,
    }


def find_kund_headers(scan):
    headers = []
    for row, col, lower, raw in scan["textIndex"]:
        if lower.startswith("kund:"):
            headers.append({"row": row, "col": col, "raw": raw, "customer": parse_customer_name(raw)})
    headers.sort(key=lambda item: (item["row"], item["col"]))
    return headers


def find_art_headers_in_window(scan, min_row, max_row, min_col, max_col):
    hits = []
    for row, col, lower, raw in scan["textIndex"]:
        if row < min_row or row > max_row or col < min_col or col > max_col:
            continue
        if lower.startswith("art nr"):
            hits.append({"row": row, "col": col, "raw": raw, "cell": a1(row, col)})
    hits.sort(key=lambda item: (item["row"], item["col"]))
    return hits


def find_date_info_for_kund(worksheet, kund_row, kund_col, search_cols=70):
    row_start = max(1, kund_row - 1)
    row_end = kund_row + 2
    col_start = kund_col
    col_end = kund_col + search_cols

    date_label = None
    date_value = None

    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            value = worksheet.cell(row=row, column=col).value
            lower = normalize_text(value)
            if lower.startswith("datum") and date_label is None:
                date_label = {"cell": a1(row, col), "value": text_value(value)}
            if is_date_like(value):
                date_value = {"cell": a1(row, col), "value": text_value(value)}

    return {"label": date_label, "value": date_value}


def find_next_kund_same_column(kund_headers, current_index):
    current = kund_headers[current_index]
    for next_index in range(current_index + 1, len(kund_headers)):
        candidate = kund_headers[next_index]
        if abs(candidate["col"] - current["col"]) <= 2 and candidate["row"] > current["row"]:
            return candidate
    return None


def find_next_kund_same_row(kund_headers, current_index):
    current = kund_headers[current_index]
    for next_index in range(current_index + 1, len(kund_headers)):
        candidate = kund_headers[next_index]
        if candidate["row"] == current["row"] and candidate["col"] > current["col"]:
            return candidate
    return None


def build_merged_ranges(worksheet):
    merged = []
    for rng in worksheet.merged_cells.ranges:
        merged.append(
            {
                "range": str(rng),
                "minRow": rng.min_row,
                "maxRow": rng.max_row,
                "minCol": rng.min_col,
                "maxCol": rng.max_col,
            }
        )
    return merged


def overlapping_merged_ranges(merged_ranges, min_row, max_row, min_col, max_col):
    hits = []
    for rng in merged_ranges:
        if rng["maxRow"] < min_row:
            continue
        if rng["minRow"] > max_row:
            continue
        if rng["maxCol"] < min_col:
            continue
        if rng["minCol"] > max_col:
            continue
        hits.append(rng["range"])
    return hits


def find_footer_row(scan, table_top_row, table_bottom_row, min_col, max_col):
    footer_keywords = ["total", "full pall", "srs-pall", "sortrena", "kolli", "platser"]
    footer_rows = []

    for row, col, lower, _raw in scan["textIndex"]:
        if row <= table_top_row + 4 or row > table_bottom_row:
            continue
        if col < min_col or col > max_col:
            continue
        if any(keyword in lower for keyword in footer_keywords):
            footer_rows.append(row)

    if not footer_rows:
        return None
    return min(footer_rows)


def collect_footer_formulas(scan, min_row, max_row, min_col, max_col):
    formulas = []
    for row, col, formula in scan["formulas"]:
        if row < min_row or row > max_row or col < min_col or col > max_col:
            continue
        formulas.append({"cell": a1(row, col), "formula": formula})
    formulas.sort(key=lambda item: item["cell"])
    return formulas


def detect_primary_sections(art_headers):
    if not art_headers:
        return {"left": None, "right": None, "additional": []}

    top_row = min(item["row"] for item in art_headers)
    top_band = [item for item in art_headers if item["row"] <= top_row + 1]
    top_band.sort(key=lambda item: item["col"])

    left = top_band[0] if top_band else art_headers[0]
    right = top_band[1] if len(top_band) > 1 else (art_headers[1] if len(art_headers) > 1 else None)

    used = {id(left)}
    if right is not None:
        used.add(id(right))

    additional = [h for h in art_headers if id(h) not in used]
    return {"left": left, "right": right, "additional": additional}


def infer_table_bounds(scan, kund_headers, index):
    current = kund_headers[index]
    next_same_col = find_next_kund_same_column(kund_headers, index)
    next_same_row = find_next_kund_same_row(kund_headers, index)

    min_row = current["row"]
    max_row = (next_same_col["row"] - 1) if next_same_col else scan["usedMaxRow"]
    min_col = current["col"]
    max_col = (next_same_row["col"] - 1) if next_same_row else scan["usedMaxCol"]

    return min_row, max_row, min_col, max_col


def analyze_single_table(worksheet, scan, merged_ranges, kund_headers, index):
    kund = kund_headers[index]
    min_row, max_row, min_col, max_col = infer_table_bounds(scan, kund_headers, index)

    art_headers = find_art_headers_in_window(
        scan,
        min_row=min_row,
        max_row=min(max_row, min_row + 25),
        min_col=min_col,
        max_col=max_col,
    )
    sections = detect_primary_sections(art_headers)
    date_info = find_date_info_for_kund(worksheet, kund["row"], kund["col"])

    data_start_row = None
    if sections["left"]:
        data_start_row = sections["left"]["row"] + 1
    elif sections["right"]:
        data_start_row = sections["right"]["row"] + 1

    footer_start_row = find_footer_row(scan, min_row, max_row, min_col, max_col)
    data_end_row = None
    if data_start_row is not None:
        if footer_start_row is not None and footer_start_row > data_start_row:
            data_end_row = footer_start_row - 1
        else:
            data_end_row = max_row

    footer_formulas = []
    if footer_start_row is not None:
        footer_formulas = collect_footer_formulas(scan, footer_start_row, max_row, min_col, max_col)

    return {
        "tableId": f"{worksheet.title}::{kund['row']}::{kund['col']}",
        "type": "special" if worksheet.title in SPECIAL_SHEETS else "standard",
        "header": {
            "kundCell": a1(kund["row"], kund["col"]),
            "kundRaw": kund["raw"],
            "customerName": kund["customer"],
            "date": date_info,
        },
        "bounds": {
            "range": range_a1(min_row, min_col, max_row, max_col),
            "minRow": min_row,
            "maxRow": max_row,
            "minCol": min_col,
            "maxCol": max_col,
        },
        "sections": {
            "left": sections["left"],
            "right": sections["right"],
            "additionalArtHeaders": sections["additional"],
        },
        "dataWindow": {
            "startRow": data_start_row,
            "endRow": data_end_row,
            "footerStartRow": footer_start_row,
        },
        "footer": {
            "formulaCount": len(footer_formulas),
            "formulas": footer_formulas,
        },
        "mergedRanges": overlapping_merged_ranges(merged_ranges, min_row, max_row, min_col, max_col),
    }


def sheet_design_note(sheet_name):
    if sheet_name == "ICA Helg":
        return (
            "Custom layout: multiple side-by-side customer blocks with additional Art nr blocks "
            "in lower rows for full-pall/SRS sections."
        )
    if sheet_name == "Axfood Pontus":
        return (
            "Custom layout: single customer block where left and right sections start on different rows "
            "(main left near row 4, right/full-pall near row 11)."
        )
    if sheet_name == "Axfood Pontus BIGGER":
        return (
            "Custom layout: larger variant of Axfood Pontus with customer/date on row 1 and lower second section."
        )
    return "Standard layout: Kund/date header, left and right Art nr sections, footer with totals/formulas."


def summarize_sheet(worksheet):
    scan = scan_sheet_cells(worksheet)
    kund_headers = find_kund_headers(scan)
    merged_ranges = build_merged_ranges(worksheet)

    tables = [
        analyze_single_table(worksheet, scan, merged_ranges, kund_headers, index)
        for index in range(len(kund_headers))
    ]

    return {
        "sheet": worksheet.title,
        "layoutType": "special" if worksheet.title in SPECIAL_SHEETS else "standard",
        "designNote": sheet_design_note(worksheet.title),
        "usedRange": range_a1(scan["usedMinRow"], scan["usedMinCol"], scan["usedMaxRow"], scan["usedMaxCol"]),
        "nonEmptyCells": len(scan["values"]),
        "formulaCells": len(scan["formulas"]),
        "mergedRangeCount": len(merged_ranges),
        "tableCount": len(tables),
        "tables": tables,
    }


def analyze_workbook(path):
    workbook = load_workbook(path, data_only=False)
    sheet_summaries = [summarize_sheet(ws) for ws in workbook.worksheets]

    return {
        "file": str(path),
        "sheetCount": len(workbook.sheetnames),
        "specialSheets": sorted(list(SPECIAL_SHEETS)),
        "sheets": sheet_summaries,
    }


def print_summary(report):
    print(f"Workbook: {report['file']}")
    print(f"Sheets: {report['sheetCount']}")

    for sheet in report["sheets"]:
        print("-" * 80)
        print(f"Sheet: {sheet['sheet']} ({sheet['layoutType']})")
        print(f"Design: {sheet['designNote']}")
        print(f"Used range: {sheet['usedRange']}")
        print(f"Cells: non-empty={sheet['nonEmptyCells']} formulas={sheet['formulaCells']}")
        print(f"Merged ranges: {sheet['mergedRangeCount']}")
        print(f"Detected customer tables: {sheet['tableCount']}")

        for idx, table in enumerate(sheet["tables"], start=1):
            customer = table["header"]["customerName"] or "<empty>"
            kund_cell = table["header"]["kundCell"]
            left = table["sections"]["left"]
            right = table["sections"]["right"]
            left_cell = left["cell"] if left else "-"
            right_cell = right["cell"] if right else "-"
            date_value = table["header"]["date"]["value"]
            date_text = date_value["value"] if date_value else "-"

            print(
                f"  [{idx}] customer={customer} | kund={kund_cell} | date={date_text} | "
                f"bounds={table['bounds']['range']}"
            )
            print(
                f"      sections: left={left_cell} right={right_cell} "
                f"additional={len(table['sections']['additionalArtHeaders'])}"
            )
            print(
                f"      data rows: {table['dataWindow']['startRow']}..{table['dataWindow']['endRow']} | "
                f"footer row: {table['dataWindow']['footerStartRow']} | "
                f"footer formulas: {table['footer']['formulaCount']}"
            )


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Analyze the standard Excel template by Kund/date header, Art nr left/right sections, "
            "and footer formulas, including special handling notes for ICA Helg and Axfood Pontus sheets."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("public/templates/Plocklist-Template.xlsx"),
        help="Path to workbook (.xlsx)",
    )
    parser.add_argument(
        "--json-out",
        type=Path,
        default=None,
        help="Optional output file for full JSON report",
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {args.workbook}")

    report = analyze_workbook(args.workbook)
    print_summary(report)

    if args.json_out:
        args.json_out.parent.mkdir(parents=True, exist_ok=True)
        args.json_out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"JSON report written to: {args.json_out}")


if __name__ == "__main__":
    main()
