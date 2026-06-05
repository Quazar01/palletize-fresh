import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SPECIAL_HINTS = {
    "ICA Helg": "Multiple side-by-side customer blocks with extra lower Art nr blocks; use additionalArtHeaderCells for secondary fill zones.",
    "Axfood Pontus": "Custom block where right/full-pall Art nr header is lower than left section.",
    "Axfood Pontus BIGGER": "Larger Axfood Pontus variant with same split logic and extended vertical area.",
}


def range_a1(min_row, min_col, max_row, max_col):
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def is_non_empty(value):
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def detect_section_end_col(worksheet, start_row, start_col, max_col, rows_to_scan=3, default_width=6):
    hit_cols = []
    for row in range(start_row, start_row + rows_to_scan):
        for col in range(start_col, max_col + 1):
            if is_non_empty(worksheet.cell(row=row, column=col).value):
                hit_cols.append(col)

    if hit_cols:
        return max(hit_cols)

    return min(max_col, start_col + default_width)


def compute_section_ranges(worksheet, table):
    bounds = table["bounds"]
    left = table["sections"]["left"]
    right = table["sections"]["right"]
    data = table["dataWindow"]

    start_row = data["startRow"]
    end_row = data["endRow"] if data["endRow"] is not None else bounds["maxRow"]

    left_range = None
    right_range = None

    if left and start_row is not None:
        if right:
            left_max_col = max(left["col"], right["col"] - 2)
        else:
            left_max_col = min(bounds["maxCol"], left["col"] + 10)

        left_end_col = detect_section_end_col(
            worksheet=worksheet,
            start_row=left["row"],
            start_col=left["col"],
            max_col=left_max_col,
            rows_to_scan=3,
            default_width=5,
        )
        left_range = range_a1(start_row, left["col"], end_row, left_end_col)

    if right and start_row is not None:
        right_end_col = detect_section_end_col(
            worksheet=worksheet,
            start_row=right["row"],
            start_col=right["col"],
            max_col=bounds["maxCol"],
            rows_to_scan=3,
            default_width=6,
        )
        right_range = range_a1(start_row, right["col"], end_row, right_end_col)

    footer_start = data["footerStartRow"]
    footer_range = None
    if footer_start is not None:
        footer_range = range_a1(footer_start, bounds["minCol"], bounds["maxRow"], bounds["maxCol"])

    return {
        "dataStartRow": start_row,
        "dataEndRow": end_row,
        "footerStartRow": footer_start,
        "leftDataRange": left_range,
        "rightDataRange": right_range,
        "footerRange": footer_range,
    }


def build_table_mapping(worksheet, sheet, table, table_index):
    customer_name = table["header"].get("customerName") or ""
    date_value = table["header"]["date"].get("value")
    date_label = table["header"]["date"].get("label")
    section_ranges = compute_section_ranges(worksheet, table)

    left = table["sections"].get("left")
    right = table["sections"].get("right")

    return {
        "tableKey": f"{sheet['sheet']}::{table_index}",
        "tableId": table["tableId"],
        "sheet": sheet["sheet"],
        "layoutType": sheet["layoutType"],
        "customerName": customer_name,
        "header": {
            "kundCell": table["header"]["kundCell"],
            "kundRaw": table["header"]["kundRaw"],
            "dateLabelCell": date_label["cell"] if date_label else None,
            "dateValueCell": date_value["cell"] if date_value else None,
            "dateValue": date_value["value"] if date_value else None,
        },
        "bounds": table["bounds"],
        "anchors": {
            "leftArtHeaderCell": left["cell"] if left else None,
            "rightArtHeaderCell": right["cell"] if right else None,
            "additionalArtHeaderCells": [item["cell"] for item in table["sections"].get("additionalArtHeaders", [])],
        },
        "ranges": section_ranges,
        "footer": {
            "formulaCount": table["footer"]["formulaCount"],
            "formulaCells": [item["cell"] for item in table["footer"]["formulas"]],
        },
        "mergedRanges": table["mergedRanges"],
        "specialHint": SPECIAL_HINTS.get(sheet["sheet"]),
    }


def generate_mappings(report, workbook):
    sheet_mappings = []
    table_mappings = []

    for sheet in report["sheets"]:
        worksheet = workbook[sheet["sheet"]]
        current_tables = []
        for table_index, table in enumerate(sheet["tables"], start=1):
            mapping = build_table_mapping(worksheet, sheet, table, table_index)
            table_mappings.append(mapping)
            current_tables.append(mapping)

        sheet_mappings.append(
            {
                "sheet": sheet["sheet"],
                "layoutType": sheet["layoutType"],
                "tableCount": len(current_tables),
                "tables": current_tables,
            }
        )

    return {
        "sourceWorkbook": report["file"],
        "sheetCount": report["sheetCount"],
        "specialSheets": report.get("specialSheets", []),
        "mappingVersion": 1,
        "sheets": sheet_mappings,
        "tables": table_mappings,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Generate strict per-sheet/per-table mapping config from template analysis JSON."
    )
    parser.add_argument(
        "--analysis-json",
        type=Path,
        default=Path("tmp/template-report-rigorous.json"),
        help="Input analysis JSON generated by scripts/analyze_template.py",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("tmp/template-table-mappings.json"),
        help="Output mapping JSON",
    )
    args = parser.parse_args()

    if not args.analysis_json.exists():
        raise FileNotFoundError(f"Analysis JSON not found: {args.analysis_json}")

    report = json.loads(args.analysis_json.read_text(encoding="utf-8"))
    source_workbook_value = report.get("sourceWorkbook") or report.get("file")
    if not source_workbook_value:
        raise KeyError("Analysis JSON must contain either 'sourceWorkbook' or 'file'.")

    source_workbook = Path(source_workbook_value)
    if not source_workbook.exists():
        raise FileNotFoundError(
            f"Workbook from analysis JSON not found: {source_workbook}. "
            f"Run command from project root or use a report with a valid relative workbook path."
        )

    workbook = load_workbook(source_workbook, data_only=False)
    mappings = generate_mappings(report, workbook)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(mappings, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Generated mapping file: {args.out}")
    print(f"Sheets: {mappings['sheetCount']}")
    print(f"Tables: {len(mappings['tables'])}")


if __name__ == "__main__":
    main()
